import openpyxl

def excel_data(path, sheetname):
    final_list = []
    
    workbook = openpyxl.load_workbook(path) 
    sheet = workbook[sheetname]
    rows = sheet.max_row 
    col = sheet.max_column
    
    for r in range(2, rows + 1):
        row_list = []
        for c in range(1, col + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
        
    return final_list